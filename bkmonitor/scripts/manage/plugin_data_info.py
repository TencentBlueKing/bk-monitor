from collections import defaultdict
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from metadata.models import (
    AccessVMRecord,
    ClusterInfo,
    DataSource,
    DataSourceResultTable,
    ResultTable,
    ResultTableOption,
)
from metadata.models.constants import DataIdCreatedFromSystem
from monitor_web.commons.data_access import PluginDataAccessor
from monitor_web.models.plugin import CollectorPluginMeta
from monitor_web.plugin.constant import PluginType

"""
1. 普通插件
2. 进程/日志关键字/snmp_trap插件
3. k8s插件
"""

operator = "admin"


class TableDataInfo(BaseModel):
    result_table_id: str
    enable_blacklist: bool
    is_split_measurement: bool
    vm_result_table_id: str
    vm_cluster_id: int
    vm_cluster_name: str


class PluginDataInfo(BaseModel):
    plugin_id: str
    plugin_type: str
    bk_biz_id: int
    is_global: bool
    data_id: int
    data_name: str
    etl_config: str
    register_to_bkbase: bool
    table_infos: list[TableDataInfo]


class PluginInfoJson(BaseModel):
    plugin_infos: list[PluginDataInfo]


def get_table_info_by_data_id(
    data_id: int, cluster_name_map: dict[int, str]
) -> tuple[list[TableDataInfo], str, str, bool]:
    """获取数据源对应的结果表信息

    Args:
        data_id: 数据源ID
        cluster_name_map: 集群名称映射
    Returns:
        table_infos: 结果表信息
        data_name: 数据源名称
        etl_config: 数据源ETL配置
        register_to_bkbase: 是否注册到BKBase
    """
    data_source = DataSource.objects.get(bk_data_id=data_id)
    dsrt = DataSourceResultTable.objects.filter(bk_data_id=data_id)
    rt_qs = ResultTable.objects.filter(table_id__in=[dsrt.table_id for dsrt in dsrt], is_enable=True)
    table_ids = list(rt_qs.values_list("table_id", flat=True))
    table_id_vm_config: dict[str, AccessVMRecord] = {
        vm.result_table_id: vm for vm in AccessVMRecord.objects.filter(result_table_id__in=table_ids)
    }

    rt_option_qs = ResultTableOption.objects.filter(
        table_id__in=table_ids,
        name__in=[ResultTableOption.OPTION_ENABLE_FIELD_BLACK_LIST, ResultTableOption.OPTION_IS_SPLIT_MEASUREMENT],
    )
    rt_options = defaultdict(dict)
    for option in rt_option_qs:
        rt_options[option.table_id][option.name] = option.get_value()

    result: list[TableDataInfo] = []
    for rt in rt_qs:
        table_id: str = rt.table_id
        access_vm_record = table_id_vm_config.get(table_id)
        vm_cluster_name = ""
        vm_cluster_id = 0
        if access_vm_record:
            if access_vm_record.vm_cluster_id:
                vm_cluster_id = access_vm_record.vm_cluster_id
                vm_cluster_name = cluster_name_map.get(vm_cluster_id, "")

        result.append(
            TableDataInfo(
                result_table_id=table_id,
                vm_result_table_id=access_vm_record.vm_result_table_id if access_vm_record else "",
                vm_cluster_id=access_vm_record.vm_cluster_id
                if access_vm_record and access_vm_record.vm_cluster_id
                else 0,
                vm_cluster_name=vm_cluster_name,
                enable_blacklist=rt_options[table_id].get(ResultTableOption.OPTION_ENABLE_FIELD_BLACK_LIST, False),
                is_split_measurement=rt_options[table_id].get(ResultTableOption.OPTION_IS_SPLIT_MEASUREMENT, False),
            )
        )
    return (
        result,
        data_source.data_name,
        data_source.etl_config,
        data_source.created_from == DataIdCreatedFromSystem.BKDATA.value,
    )


def get_plugin_infos(
    plugin_types: list[str] | None = None, plugin_ids: list[str] | None = None
) -> list[PluginDataInfo]:
    plugins = CollectorPluginMeta.objects.all()

    if plugin_types:
        plugins = plugins.filter(plugin_type__in=plugin_types)
    if plugin_ids:
        plugins = plugins.filter(plugin_id__in=plugin_ids)

    # 获取集群名称映射
    cluster_name_map: dict[int, str] = dict(ClusterInfo.objects.values_list("cluster_id", "cluster_name"))

    result: list[PluginDataInfo] = []
    for plugin in plugins:
        if plugin.plugin_type in [
            PluginType.SCRIPT,
            PluginType.DATADOG,
            PluginType.EXPORTER,
            PluginType.JMX,
            PluginType.PUSHGATEWAY,
        ]:
            manager = PluginDataAccessor(plugin.current_version, operator=operator)

            bk_data_id = manager.get_data_id()
            if bk_data_id:
                table_infos, data_name, etl_config, register_to_bkbase = get_table_info_by_data_id(
                    bk_data_id, cluster_name_map
                )
            else:
                table_infos, data_name, etl_config, register_to_bkbase = [], "", "", False

            result.append(
                PluginDataInfo(
                    plugin_id=plugin.plugin_id,
                    plugin_type=plugin.plugin_type,
                    bk_biz_id=plugin.bk_biz_id,
                    is_global=plugin.is_global,
                    data_id=bk_data_id if bk_data_id else 0,
                    data_name=data_name,
                    etl_config=etl_config,
                    register_to_bkbase=register_to_bkbase,
                    table_infos=table_infos,
                )
            )
        elif plugin.plugin_type in [PluginType.LOG, PluginType.SNMP_TRAP]:
            pass

    return result


def save_to_json(plugin_infos: list[PluginDataInfo], output_file: str = "plugin_data_info.json") -> None:
    """保存插件数据信息到JSON文件

    Args:
        plugin_infos: 插件数据信息
    """
    plugin_info_json = PluginInfoJson(plugin_infos=plugin_infos)
    with open(output_file, "w") as f:
        f.write(plugin_info_json.model_dump_json(indent=2))


def json_to_excel(
    plugin_infos: list[dict[str, Any]] | list[PluginDataInfo], output_file: str = "plugin_data_info.xlsx"
) -> None:
    """将插件数据信息转换为Excel文件

    Args:
        plugin_infos: 插件数据信息
    """
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "plugin_data_info"

    headers = [
        "插件ID",
        "插件类型",
        "业务ID",
        "全局插件",
        "数据ID",
        "数据名称",
        "清洗配置",
        "是否注册到bkbase",
        "结果表ID",
        "是否自动发现",
        "单指标单表",
        "vm结果表",
        "vm集群ID",
        "vm集群名称",
    ]
    worksheet.append(headers)

    normalized_plugin_infos = [
        raw_plugin_info
        if isinstance(raw_plugin_info, PluginDataInfo)
        else PluginDataInfo.model_validate(raw_plugin_info)
        for raw_plugin_info in plugin_infos
    ]
    sorted_plugin_infos = sorted(normalized_plugin_infos, key=lambda item: (item.data_id == 0, item.bk_biz_id))

    for plugin_info in sorted_plugin_infos:
        # 将嵌套的结果表信息摊平成多行，便于在 Excel 中筛选和统计。
        table_infos: list[TableDataInfo | None] = [*plugin_info.table_infos] if plugin_info.table_infos else [None]
        start_row = worksheet.max_row + 1
        for table_info in table_infos:
            worksheet.append(
                [
                    plugin_info.plugin_id,
                    plugin_info.plugin_type,
                    plugin_info.bk_biz_id,
                    plugin_info.is_global,
                    plugin_info.data_id,
                    plugin_info.data_name,
                    plugin_info.etl_config,
                    plugin_info.register_to_bkbase,
                    table_info.result_table_id if table_info else "",
                    table_info.enable_blacklist if table_info else "",
                    table_info.is_split_measurement if table_info else "",
                    table_info.vm_result_table_id if table_info else "",
                    table_info.vm_cluster_id if table_info else "",
                    table_info.vm_cluster_name if table_info else "",
                ]
            )
        end_row = worksheet.max_row

        if end_row > start_row:
            for column_index in range(1, 9):
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=column_index,
                    end_row=end_row,
                    end_column=column_index,
                )
                worksheet.cell(row=start_row, column=column_index).alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        column_letter = get_column_letter(column_index)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
