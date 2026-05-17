"""统计 etl_config 为 bk_exporter / bk_standard 的插件数据在告警策略和 Grafana 仪表盘中的使用情况。

使用方式（在 Django shell 或管理脚本中）：
    from scripts.manage.plugin_usage_stats import run
    run(output_file="plugin_usage_stats.xlsx")
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from pydantic import BaseModel

from bk_dataview.models import Dashboard, Org
from bkmonitor.models import QueryConfigModel, StrategyModel
from metadata.models import DataSource, DataSourceResultTable, ResultTable

TARGET_ETL_CONFIGS = ("bk_exporter", "bk_standard")

# 判定为"已聚合主机维度"的字段名，包含任一即视为已按主机聚合
HOST_DIMENSIONS = {"ip", "bk_target_ip"}


# ---------------------------------------------------------------------------
# Pydantic 数据结构
# ---------------------------------------------------------------------------


class ResultTableInfo(BaseModel):
    """目标结果表及其关联的数据源信息"""

    table_id: str
    data_label: str
    bk_data_id: int
    data_name: str
    etl_config: str


class StrategyHit(BaseModel):
    """命中的告警策略记录"""

    strategy_id: int
    strategy_name: str
    bk_biz_id: int
    is_enabled: bool
    data_source_label: str
    data_type_label: str
    metric_id: str
    matched_keywords: list[str]
    agg_dimensions: list[str]
    has_host_dimension: bool


class DashboardHit(BaseModel):
    """命中的仪表盘记录"""

    bk_biz_id: str
    dashboard_title: str
    dashboard_uid: str
    matched_keywords: list[str]
    has_host_dimension: bool


# ---------------------------------------------------------------------------
# Step 1: 收集目标结果表
# ---------------------------------------------------------------------------


def collect_target_tables() -> tuple[dict[str, ResultTableInfo], set[str], set[str]]:
    """收集 etl_config 为 bk_exporter / bk_standard 的数据源关联的结果表。

    Returns:
        rt_map: table_id -> ResultTableInfo
        table_id_set: 所有目标 table_id 集合
        data_label_set: 所有非空 data_label 集合
    """
    ds_qs = DataSource.objects.filter(etl_config__in=TARGET_ETL_CONFIGS).values("bk_data_id", "data_name", "etl_config")
    ds_map: dict[int, dict[str, Any]] = {row["bk_data_id"]: row for row in ds_qs}

    dsrt_qs = DataSourceResultTable.objects.filter(bk_data_id__in=list(ds_map.keys())).values("bk_data_id", "table_id")
    table_id_to_data_id: dict[str, int] = {row["table_id"]: row["bk_data_id"] for row in dsrt_qs}

    rt_qs = ResultTable.objects.filter(table_id__in=list(table_id_to_data_id.keys())).values("table_id", "data_label")

    rt_map: dict[str, ResultTableInfo] = {}
    data_label_set: set[str] = set()

    for rt in rt_qs:
        table_id: str = rt["table_id"]
        data_label: str = rt["data_label"] or ""
        bk_data_id = table_id_to_data_id[table_id]
        ds_info = ds_map[bk_data_id]

        rt_map[table_id] = ResultTableInfo(
            table_id=table_id,
            data_label=data_label,
            bk_data_id=bk_data_id,
            data_name=ds_info["data_name"],
            etl_config=ds_info["etl_config"],
        )
        if data_label:
            data_label_set.add(data_label)

    table_id_set = set(rt_map.keys())
    print(f"[collect] 目标数据源: {len(ds_map)}, 结果表: {len(rt_map)}, 非空 data_label: {len(data_label_set)}")
    return rt_map, table_id_set, data_label_set


# ---------------------------------------------------------------------------
# Step 2: 构建搜索关键词与正则
# ---------------------------------------------------------------------------


def build_search_patterns(table_id_set: set[str], data_label_set: set[str]) -> tuple[set[str], re.Pattern[str]]:
    """为目标 table_id / data_label 构建搜索关键词集合和统一正则。

    对每个 table_id 生成两种变体：
      - 原始: exporter_xxx.group
      - PromQL 冒号: exporter_xxx:group

    Returns:
        all_keywords: 所有关键词（用于逐个匹配定位命中项）
        combined_pattern: 编译后的正则，任一关键词命中即 match
    """
    all_keywords: set[str] = set()

    for table_id in table_id_set:
        all_keywords.add(table_id)
        colon_variant = table_id.replace(".", ":")
        if colon_variant != table_id:
            all_keywords.add(colon_variant)

    all_keywords.update(data_label_set)

    escaped = [re.escape(kw) for kw in sorted(all_keywords, key=len, reverse=True)]
    combined_pattern = re.compile("|".join(escaped))

    print(f"[build] 搜索关键词总数: {len(all_keywords)}")
    return all_keywords, combined_pattern


# ---------------------------------------------------------------------------
# Step 3: 查询告警策略
# ---------------------------------------------------------------------------


def _extract_agg_dimensions(config: dict, data_source_label: str) -> list[str]:
    """从 QueryConfig.config 中提取聚合维度列表。

    非 PromQL 策略直接读取 agg_dimension 字段；
    PromQL 策略尝试从 promql 文本中粗提取 by(...) 子句里的维度名。
    """
    if data_source_label != "prometheus":
        return config.get("agg_dimension", []) or []

    promql: str = config.get("promql", "")
    if not promql:
        return []

    # 粗提取所有 by(...) / by (...) 子句中的标识符
    dims: set[str] = set()
    for m in re.finditer(r"\bby\s*\(([^)]*)\)", promql, re.IGNORECASE):
        for dim in re.findall(r"[a-zA-Z_]\w*", m.group(1)):
            dims.add(dim)
    return sorted(dims)


def _has_host_dimension(agg_dimensions: list[str]) -> bool:
    """判断聚合维度中是否包含主机维度。"""
    return bool(HOST_DIMENSIONS & set(agg_dimensions))


def query_strategies(
    table_id_set: set[str],
    data_label_set: set[str],
    all_keywords: set[str],
    combined_pattern: re.Pattern[str],
) -> list[StrategyHit]:
    """查询使用了目标结果表且未按主机维度聚合的告警策略。

    - 非 PromQL: 用 JSONField 精确匹配 config.result_table_id / config.data_label
    - PromQL: 对 config 文本做正则粗筛
    - 过滤: 仅保留 agg_dimension 中不包含 ip / bk_target_ip 的记录
    """
    from django.db.models import Q

    # --- 非 PromQL 精确查询 ---
    q_filters = Q(config__result_table_id__in=list(table_id_set))
    if data_label_set:
        q_filters |= Q(config__data_label__in=list(data_label_set))

    non_promql_qc_ids: set[int] = set(
        QueryConfigModel.objects.exclude(data_source_label="prometheus").filter(q_filters).values_list("id", flat=True)
    )

    # --- PromQL 正则粗筛 ---
    promql_qc_ids: set[int] = set()
    promql_qs = QueryConfigModel.objects.filter(data_source_label="prometheus").values("id", "config")
    for row in promql_qs.iterator():
        config_text = json.dumps(row["config"]) if isinstance(row["config"], dict) else str(row["config"])
        if combined_pattern.search(config_text):
            promql_qc_ids.add(row["id"])

    matched_qc_ids = non_promql_qc_ids | promql_qc_ids
    print(
        f"[strategy] 非PromQL命中: {len(non_promql_qc_ids)}, "
        f"PromQL命中: {len(promql_qc_ids)}, 合计: {len(matched_qc_ids)}"
    )

    if not matched_qc_ids:
        return []

    # 获取 QueryConfig 详情，确定每条命中了哪些关键词，并提取聚合维度
    qc_rows = QueryConfigModel.objects.filter(id__in=matched_qc_ids).values(
        "id", "strategy_id", "data_source_label", "data_type_label", "metric_id", "config"
    )

    strategy_ids: set[int] = set()
    qc_details: list[dict[str, Any]] = []
    skipped_has_host_dim = 0

    for row in qc_rows:
        config = row["config"] if isinstance(row["config"], dict) else {}
        config_text = json.dumps(row["config"]) if isinstance(row["config"], dict) else str(row["config"])
        matched = sorted({m.group() for m in combined_pattern.finditer(config_text)} & all_keywords)

        agg_dims = _extract_agg_dimensions(config, row["data_source_label"])
        has_host_dim = _has_host_dimension(agg_dims)

        if has_host_dim:
            skipped_has_host_dim += 1
            continue

        strategy_ids.add(row["strategy_id"])
        qc_details.append({**row, "matched_keywords": matched, "agg_dimensions": agg_dims})

    print(f"[strategy] 已聚合主机维度(跳过): {skipped_has_host_dim}, 缺少主机维度(保留): {len(qc_details)}")

    strategy_map: dict[int, dict[str, Any]] = {
        s["id"]: s
        for s in StrategyModel.objects.filter(id__in=strategy_ids).values("id", "name", "bk_biz_id", "is_enabled")
    }

    results: list[StrategyHit] = []
    for qc in qc_details:
        strategy = strategy_map.get(qc["strategy_id"])
        if not strategy:
            continue
        results.append(
            StrategyHit(
                strategy_id=strategy["id"],
                strategy_name=strategy["name"],
                bk_biz_id=strategy["bk_biz_id"],
                is_enabled=strategy["is_enabled"],
                data_source_label=qc["data_source_label"],
                data_type_label=qc["data_type_label"],
                metric_id=qc["metric_id"],
                matched_keywords=qc["matched_keywords"],
                agg_dimensions=qc["agg_dimensions"],
                has_host_dimension=False,
            )
        )

    results.sort(key=lambda h: (h.bk_biz_id, h.strategy_id))
    print(f"[strategy] 最终命中策略记录数: {len(results)}")
    return results


# ---------------------------------------------------------------------------
# Step 4: 查询 Grafana 仪表盘
# ---------------------------------------------------------------------------


# 粗筛仪表盘 JSON 中是否包含主机维度关键词（group_by 中的 "ip" 或 "bk_target_ip"）
# 匹配 JSON 中 "ip" 或 "bk_target_ip" 作为独立字符串值的情况
_DASHBOARD_HOST_DIM_PATTERN = re.compile(r'"(?:ip|bk_target_ip)"')


def query_dashboards(
    all_keywords: set[str],
    combined_pattern: re.Pattern[str],
) -> list[DashboardHit]:
    """对所有非文件夹 Dashboard 的 JSON 文本做正则粗筛，找出使用了目标结果表的仪表盘。

    同时粗筛 JSON 中是否包含主机维度关键词，仅保留不包含的记录。
    """
    org_biz_map: dict[int, str] = dict(Org.objects.values_list("id", "name"))

    results: list[DashboardHit] = []
    dashboard_qs = Dashboard.objects.filter(is_folder=0).values("org_id", "title", "uid", "data")

    total = 0
    hit = 0
    skipped_has_host_dim = 0
    for row in dashboard_qs.iterator():
        total += 1
        data_text: str = row["data"] or ""
        if not data_text:
            continue

        matched = sorted({m.group() for m in combined_pattern.finditer(data_text)} & all_keywords)
        if not matched:
            continue

        hit += 1
        has_host_dim = bool(_DASHBOARD_HOST_DIM_PATTERN.search(data_text))
        if has_host_dim:
            skipped_has_host_dim += 1
            continue

        bk_biz_id = org_biz_map.get(row["org_id"], str(row["org_id"]))
        results.append(
            DashboardHit(
                bk_biz_id=bk_biz_id,
                dashboard_title=row["title"],
                dashboard_uid=row["uid"] or "",
                matched_keywords=matched,
                has_host_dimension=False,
            )
        )

    results.sort(key=lambda h: (h.bk_biz_id, h.dashboard_title))
    print(
        f"[dashboard] 扫描仪表盘: {total}, 命中: {hit}, "
        f"已聚合主机维度(跳过): {skipped_has_host_dim}, 缺少主机维度(保留): {len(results)}"
    )
    return results


# ---------------------------------------------------------------------------
# Step 5: 输出 Excel
# ---------------------------------------------------------------------------


def save_to_excel(
    rt_map: dict[str, ResultTableInfo],
    strategy_hits: list[StrategyHit],
    dashboard_hits: list[DashboardHit],
    output_file: str = "plugin_usage_stats.xlsx",
) -> None:
    """将统计结果输出为 Excel 文件。

    Args:
        rt_map: 目标结果表映射
        strategy_hits: 命中的策略列表
        dashboard_hits: 命中的仪表盘列表
        output_file: 输出文件路径
    """
    from typing import cast

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    def _auto_column_width(worksheet: Worksheet) -> None:
        for col_idx, col_cells in enumerate(worksheet.iter_cols(), start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 80)

    workbook = Workbook()

    # --- Sheet 0: 目标结果表 ---
    ws_rt = cast(Worksheet, workbook.active)
    ws_rt.title = "目标结果表"
    ws_rt.append(["结果表ID", "数据标签", "数据ID", "数据名称", "清洗配置"])
    for rt_info in sorted(rt_map.values(), key=lambda r: r.table_id):
        ws_rt.append([rt_info.table_id, rt_info.data_label, rt_info.bk_data_id, rt_info.data_name, rt_info.etl_config])
    ws_rt.freeze_panes = "A2"
    ws_rt.auto_filter.ref = ws_rt.dimensions
    _auto_column_width(ws_rt)

    # --- Sheet 1: 告警策略（缺少主机维度聚合的） ---
    ws_strategy = cast(Worksheet, workbook.create_sheet("告警策略"))
    ws_strategy.append(
        [
            "策略ID",
            "策略名称",
            "业务ID",
            "是否启用",
            "数据来源",
            "数据类型",
            "指标ID",
            "聚合维度",
            "命中关键词",
        ]
    )
    for hit in strategy_hits:
        ws_strategy.append(
            [
                hit.strategy_id,
                hit.strategy_name,
                hit.bk_biz_id,
                hit.is_enabled,
                hit.data_source_label,
                hit.data_type_label,
                hit.metric_id,
                ", ".join(hit.agg_dimensions),
                ", ".join(hit.matched_keywords),
            ]
        )
    ws_strategy.freeze_panes = "A2"
    ws_strategy.auto_filter.ref = ws_strategy.dimensions
    _auto_column_width(ws_strategy)

    # --- Sheet 2: Grafana 仪表盘 ---
    ws_dashboard = cast(Worksheet, workbook.create_sheet("Grafana仪表盘"))
    ws_dashboard.append(["业务ID", "仪表盘标题", "仪表盘UID", "命中关键词"])
    for hit in dashboard_hits:
        ws_dashboard.append(
            [
                hit.bk_biz_id,
                hit.dashboard_title,
                hit.dashboard_uid,
                ", ".join(hit.matched_keywords),
            ]
        )
    ws_dashboard.freeze_panes = "A2"
    ws_dashboard.auto_filter.ref = ws_dashboard.dimensions
    _auto_column_width(ws_dashboard)

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    print(f"[export] 已保存到 {output_path.resolve()}")


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------


def run(output_file: str | None = "plugin_usage_stats.xlsx") -> None:
    """执行完整的统计流程。

    Args:
        output_file: Excel 输出路径。为 None 时只打印汇总，不生成文件。
    """
    rt_map, table_id_set, data_label_set = collect_target_tables()
    if not rt_map:
        print("未找到 etl_config 为 bk_exporter/bk_standard 的结果表，退出。")
        return

    all_keywords, combined_pattern = build_search_patterns(table_id_set, data_label_set)

    strategy_hits = query_strategies(table_id_set, data_label_set, all_keywords, combined_pattern)
    dashboard_hits = query_dashboards(all_keywords, combined_pattern)

    print("\n===== 汇总 =====")
    print(f"目标结果表: {len(rt_map)}")
    print(f"命中告警策略: {len(strategy_hits)}")
    print(f"命中仪表盘: {len(dashboard_hits)}")

    if output_file:
        save_to_excel(rt_map, strategy_hits, dashboard_hits, output_file)
